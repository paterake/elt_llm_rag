"""Document preprocessor framework for transforming files before ingestion.

This module provides a framework for running preprocessors on files before
they are ingested into the RAG system. Preprocessors can transform files
(e.g., XML to Markdown) to improve embedding quality.
"""

from __future__ import annotations

import csv
import logging
import re
from abc import ABC, abstractmethod
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


@dataclass
class PreprocessorResult:
    """Result of preprocessing.

    Attributes:
        original_file: Path to the original input file.
        output_files: List of paths to generated output files.
        success: Whether preprocessing succeeded.
        message: Optional message describing the result.
        section_collection_map: For split-mode preprocessors only — maps each
            output file path to the ChromaDB collection it should be ingested
            into. None for standard single-output preprocessors.
    """
    original_file: str
    output_files: List[str]
    success: bool = True
    message: Optional[str] = None
    section_collection_map: Optional[Dict[str, str]] = None  # file_path → collection_name


class BasePreprocessor(ABC):
    """Abstract base class for document preprocessors.
    
    Subclasses must implement the preprocess() method.
    """
    
    @abstractmethod
    def preprocess(self, input_file: str, output_path: str, **kwargs: Any) -> PreprocessorResult:
        """Preprocess a file and save the output.
        
        Args:
            input_file: Path to the input file.
            output_path: Base path for output file(s).
            **kwargs: Additional format-specific arguments.
            
        Returns:
            PreprocessorResult with output file paths.
        """
        pass


class LeanIXPreprocessor(BasePreprocessor):
    """Preprocessor for LeanIX draw.io XML exports.

    Extracts assets and relationships from LeanIX diagrams and outputs
    structured Markdown suitable for RAG embedding.

    Supports three output modes:
    - ``'markdown'`` / ``'md'`` / ``'both'``: single-file output (original behaviour).
    - ``'split'``: generates one Markdown file per logical section (overview,
      one per domain group, additional entities, relationships) and populates
      :attr:`PreprocessorResult.section_collection_map` so the ingestion
      pipeline can load each file into its own ChromaDB collection.
    """

    def __init__(
        self,
        output_format: str = "markdown",
        collection_prefix: Optional[str] = None,
        model_name: str = "Enterprise Conceptual Data Model",
        org_name: str = "the organisation",
    ):
        """Initialise the LeanIX preprocessor.

        Args:
            output_format: ``'markdown'``, ``'json'``, ``'both'``, or ``'split'``.
            collection_prefix: Required when ``output_format='split'``. Each
                section file is loaded into ``{collection_prefix}_{section_key}``.
            model_name: Human-readable name of the conceptual model, used in
                generated Markdown prose (e.g. "FA Enterprise Conceptual Data Model").
            org_name: Organisation name used in generated Markdown prose
                (e.g. "The Football Association").
        """
        self.output_format = output_format
        self.collection_prefix = collection_prefix
        self.model_name = model_name
        self.org_name = org_name

    def preprocess(self, input_file: str, output_path: str, **kwargs: Any) -> PreprocessorResult:
        """Preprocess a LeanIX XML file.

        Args:
            input_file: Path to the LeanIX XML file.
            output_path: Base path for output file(s). For split mode this is
                used as the stem of the sections sub-directory.
            **kwargs: Additional arguments (currently unused).

        Returns:
            PreprocessorResult with paths to generated files, and
            ``section_collection_map`` populated when in split mode.
        """
        from .doc_leanix_parser import LeanIXExtractor

        input_path = Path(input_file).expanduser().resolve()
        output_path_obj = Path(output_path).expanduser().resolve()
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)

        logger.info("Preprocessing LeanIX file: %s", input_path)

        try:
            extractor = LeanIXExtractor(
                str(input_path),
                model_name=self.model_name,
                org_name=self.org_name,
            )
            extractor.parse_xml()
            extractor.extract_all()

            # ── Split mode: one file per domain section + relationships ────────
            if self.output_format == "split":
                section_dir = output_path_obj.parent / f"{output_path_obj.stem}_sections"
                section_file_map = extractor.save_sections(str(section_dir))

                section_collection_map: Optional[Dict[str, str]] = None
                if self.collection_prefix:
                    section_collection_map = {
                        file_path: f"{self.collection_prefix}_{section_key}"
                        for section_key, file_path in section_file_map.items()
                    }
                    logger.info(
                        "Split into %d sections → collections: %s",
                        len(section_file_map),
                        list(section_collection_map.values()),
                    )

                return PreprocessorResult(
                    original_file=str(input_path),
                    output_files=list(section_file_map.values()),
                    success=True,
                    message=(
                        f"Split into {len(section_file_map)} sections from "
                        f"{len(extractor.assets)} assets and "
                        f"{len(extractor.relationships)} relationships"
                    ),
                    section_collection_map=section_collection_map,
                )

            # ── Standard single-file modes (markdown / json / both) ───────────
            output_files: List[str] = []

            if self.output_format in ("markdown", "md", "both"):
                md_path = output_path_obj.with_suffix(".md")
                with open(md_path, "w", encoding="utf-8") as f:
                    f.write(extractor.to_markdown())
                output_files.append(str(md_path))
                logger.info("Generated Markdown: %s", md_path)

            if self.output_format in ("json", "both"):
                json_path = output_path_obj.with_suffix(".json")
                with open(json_path, "w", encoding="utf-8") as f:
                    f.write(extractor.to_json())
                output_files.append(str(json_path))
                logger.info("Generated JSON: %s", json_path)

            return PreprocessorResult(
                original_file=str(input_path),
                output_files=output_files,
                success=True,
                message=(
                    f"Extracted {len(extractor.assets)} assets and "
                    f"{len(extractor.relationships)} relationships"
                ),
            )

        except Exception as e:
            logger.error("Failed to preprocess LeanIX file: %s", e)
            return PreprocessorResult(
                original_file=str(input_path),
                output_files=[],
                success=False,
                message=str(e),
            )


class LeanIXInventoryPreprocessor(BasePreprocessor):
    """Preprocessor for the LeanIX full-inventory Excel export.

    Reads the LeanIX inventory Excel file (exported from LeanIX → Inventory →
    Export) and generates one Markdown file per fact-sheet type, then maps each
    file to its own ChromaDB collection via ``section_collection_map``.

    This mirrors the split-mode behaviour of ``LeanIXPreprocessor`` so that:
    - ``ingest_fa_ea_leanix.yaml`` → ``fa_leanix_cm_*`` (conceptual model, from XML)
    - ``ingest_fa_leanix_inventory.yaml`` → ``fa_leanix_inv_*`` (inventory, from Excel)

    Collections produced (prefix = collection_prefix, default 'leanix_inv'):
        {prefix}_dataobject    — DataObject fact sheets with definitions
        {prefix}_interface     — Interface fact sheets (data flows)
        {prefix}_application   — Application fact sheets
        {prefix}_capability    — BusinessCapability fact sheets
        {prefix}_organization  — Organization fact sheets
        {prefix}_itcomponent   — IT Component fact sheets
        {prefix}_provider      — Provider fact sheets
        {prefix}_objective     — Objective fact sheets

    Excel format expected (LeanIX standard inventory export):
        Columns: id, type, name, displayName, description, level, status, lxState
        Sheet:   first non-'ReadMe' sheet (name includes export timestamp)
    """

    # Maps LeanIX type → (collection_suffix, human_label)
    _TYPE_MAP: Dict[str, Tuple[str, str]] = {
        "DataObject":        ("dataobject",   "DataObject Inventory"),
        "Interface":         ("interface",    "Interface Inventory (Data Flows)"),
        "Application":       ("application",  "Application Inventory"),
        "BusinessCapability":("capability",   "Business Capability Inventory"),
        "Organization":      ("organization", "Organization Inventory"),
        "ITComponent":       ("itcomponent",  "IT Component Inventory"),
        "Provider":          ("provider",     "Provider Inventory"),
        "Objective":         ("objective",    "Objective Inventory"),
    }

    def __init__(
        self,
        output_format: str = "split",
        collection_prefix: Optional[str] = None,
        org_name: str = "The Organisation",
    ):
        self.output_format = output_format  # kept for interface compatibility; always split
        self.collection_prefix = collection_prefix or "leanix_inv"
        self.org_name = org_name

    def preprocess(self, input_file: str, output_path: str, **kwargs: Any) -> PreprocessorResult:
        input_path = Path(input_file).expanduser().resolve()
        output_path_obj = Path(output_path).expanduser().resolve()
        section_dir = output_path_obj.parent / f"{output_path_obj.stem}_sections"
        section_dir.mkdir(parents=True, exist_ok=True)

        logger.info("Preprocessing LeanIX inventory Excel: %s", input_path)

        try:
            rows = self._read_excel(input_path)
            logger.info("Read %d rows from Excel", len(rows))

            # Group by type
            by_type: Dict[str, List[dict]] = defaultdict(list)
            for row in rows:
                fs_type = row.get("type", "Unknown")
                if fs_type in self._TYPE_MAP:
                    by_type[fs_type].append(row)

            section_collection_map: Dict[str, str] = {}
            output_files: List[str] = []
            total = 0

            for fs_type, type_rows in sorted(by_type.items()):
                suffix, label = self._TYPE_MAP[fs_type]
                collection_name = f"{self.collection_prefix}_{suffix}"

                md_content = self._type_to_markdown(fs_type, label, type_rows)
                out_file = section_dir / f"{suffix}.md"
                out_file.write_text(md_content, encoding="utf-8")

                section_collection_map[str(out_file)] = collection_name
                output_files.append(str(out_file))
                total += len(type_rows)
                logger.info("  %s → %s (%d rows)", out_file.name, collection_name, len(type_rows))

            return PreprocessorResult(
                original_file=str(input_path),
                output_files=output_files,
                success=True,
                message=(
                    f"Split {total} fact sheets across {len(output_files)} collections "
                    f"(prefix: {self.collection_prefix})"
                ),
                section_collection_map=section_collection_map,
            )

        except Exception as e:
            logger.error("Failed to preprocess LeanIX inventory Excel: %s", e)
            return PreprocessorResult(
                original_file=str(input_path),
                output_files=[],
                success=False,
                message=str(e),
            )

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_excel(xlsx_path: Path) -> List[dict]:
        """Read the LeanIX inventory Excel export → list of row dicts.

        Finds the first non-'ReadMe' sheet (the export sheet has a
        timestamp-based name that changes with every export).
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to read LeanIX Excel exports. "
                "Add it to elt_llm_ingest dependencies: uv add openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_name = next(
            (name for name in wb.sheetnames if name.lower() != "readme"),
            wb.sheetnames[0],
        )
        ws = wb[sheet_name]

        headers: List[str] = []
        rows: List[dict] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(row)]
            else:
                if any(cell is not None for cell in row):
                    rows.append(dict(zip(headers, row)))
        return rows

    def _type_to_markdown(self, fs_type: str, label: str, rows: List[dict]) -> str:
        """Generate Markdown for a single fact-sheet type."""
        md: List[str] = []
        md.append(f"# LeanIX {label}\n\n")
        md.append(
            f"The {self.org_name} LeanIX inventory contains {len(rows)} {fs_type} fact sheets. "
            "Each entry below includes the name, LeanIX identifier, hierarchy level, "
            "and a description where recorded.\n\n"
        )
        md.append("---\n\n")

        for row in rows:
            name = (row.get("name") or row.get("displayName") or "Unknown")
            fsid = row.get("id", "")
            level = row.get("level", "")
            lx_state = row.get("lxState", "")
            description = (row.get("description") or "").strip()

            md.append(f"## {name}\n\n")
            md.append(f"**LeanIX ID**: `{fsid}`  \n")
            if level:
                md.append(f"**Level**: {level}  \n")
            if lx_state:
                md.append(f"**Quality State**: {lx_state}  \n")

            if fs_type == "Interface":
                source, target = self._parse_interface_endpoints(str(name))
                if source and target:
                    md.append(f"**Data flow**: {source} → {target}  \n")

            md.append("\n")

            if description:
                if len(description) > 800:
                    description = description[:800] + "…"
                md.append(f"{description}\n\n")
            else:
                md.append("_No description recorded in LeanIX._\n\n")

            md.append("---\n\n")

        return "".join(md)

    @staticmethod
    def _parse_interface_endpoints(name: str) -> Tuple[Optional[str], Optional[str]]:
        """Try to extract source and target from an interface name like 'A to B'."""
        match = re.match(r"^(.+?)\s+to\s+(.+?)(?:\s+LI)?$", name, re.IGNORECASE)
        if match:
            return match.group(1).strip(), match.group(2).strip()
        return None, None


class RegulatoryPDFPreprocessor(BasePreprocessor):
    """Preprocessor for structured regulatory PDF documents.

    Works well for any numbered-section regulatory PDF (legislation, association
    handbooks, standards) that follows the common pattern of:
    - Repeating section headers on every page (e.g. "8 - RULES OF THE ASSOCIATION")
    - Glossary/definitions sections using ``TERM means DEFINITION;`` or
      ``TERM  -  DEFINITION`` style

    Improves RAG retrieval quality by:

    1. **Noise stripping** — removes repeating page-level section headers and
       standard footer lines that appear on every page and inflate BM25 token
       frequencies.

    2. **Clean full-text output** — emits stripped text as a single Markdown
       file (``<stem>_clean.md``) with section headings preserved as ``##``
       markers.  Pair with a ``chunking`` override of 512 tokens so complete
       rule paragraphs stay together.

    3. **Definitions glossary output** — extracts all term-definition pairs from
       configurable page ranges and emits them as ``<stem>_definitions.md``.
       Each term gets its own ``##`` heading, giving BM25 a direct signal when
       queries mention entity names.

    Both output files are returned in ``PreprocessorResult.output_files`` and
    ingested into the same collection.

    Requires ``pypdf`` (``uv add pypdf --package elt-llm-ingest``).

    Configuration
    -------------
    All parameters are forwarded from the YAML ``preprocessor:`` block via
    ``extra_params`` — no code changes needed for new documents::

        preprocessor:
          class: "RegulatoryPDFPreprocessor"
          definitions_label: "My Document defined term"
          # Optional: pin known glossary pages for highest-confidence extraction.
          # Auto-detection runs regardless; explicit pages are tagged "explicit"
          # or "both" (found by both strategies).
          def_sections:
            - [88, 108]   # 0-based page indices, exclusive end
            - [472, 477]
          # Optional: lower/raise the auto-detect sensitivity (default 3).
          def_density_threshold: 3
    """

    # Matches repeating section headers: "N - TITLE" or "TITLE - SUBTITLE"
    # (all-caps, numbers, and punctuation only — lowercase signals body text)
    _HEADER_PATTERN = re.compile(
        r"^(?:\d+\s*[-–]\s*[A-Z\s'/&,\-]+|"
        r"[A-Z\s'/&,\-]+\s*[-–]\s*[A-Z\s'/&,\-]+)$",
        re.MULTILINE,
    )
    # Standard regulatory PDF footer / table-header noise
    _NOISE_LINES = re.compile(
        r"^(?:RETURN TO CONTENTS PAGE|DEFINITION\s+INTERPRETATION|CONTENTS|SECTION\s+PAGE)\s*$",
        re.MULTILINE,
    )

    # Defaults — override via YAML config (def_sections / definitions_label)
    # Empty list means no definition extraction unless configured.
    _DEFAULT_DEF_SECTIONS: List[Tuple[int, int]] = []
    _DEFAULT_DEFINITIONS_LABEL = "defined term"

    def __init__(
        self,
        output_format: str = "markdown",
        collection_prefix: Optional[str] = None,
        def_sections: Optional[List] = None,
        definitions_label: Optional[str] = None,
        def_density_threshold: int = 3,
    ):
        """Initialise the preprocessor.

        Args:
            output_format: Kept for interface compatibility; always produces
                two Markdown files regardless of this value.
            collection_prefix: Kept for interface compatibility (unused — both
                output files go to the same collection).
            def_sections: Optional list of ``[page_start, page_end]`` 0-indexed
                page ranges to scan for term definitions.  When supplied these
                pages are always extracted and tagged ``source: explicit``.
                Supply via YAML: ``def_sections: [[88, 108], [472, 477]]``.
            definitions_label: Short label used in the definitions Markdown
                (e.g. "FA Handbook defined term").  Defaults to "defined term".
            def_density_threshold: Minimum number of ``means``-pattern matches
                per page for auto-detection.  Pages meeting this threshold are
                tagged ``source: detected``.  Set to 0 to disable auto-detect.
                Default 3.
        """
        self.output_format = output_format
        self.collection_prefix = collection_prefix
        # YAML delivers lists-of-lists; normalise to list of tuples
        if def_sections is not None:
            self.def_sections: List[Tuple[int, int]] = [tuple(s) for s in def_sections]  # type: ignore[misc]
        else:
            self.def_sections = list(self._DEFAULT_DEF_SECTIONS)
        self.definitions_label: str = (
            definitions_label if definitions_label is not None else self._DEFAULT_DEFINITIONS_LABEL
        )
        self.def_density_threshold = def_density_threshold

    # ------------------------------------------------------------------
    # BasePreprocessor interface
    # ------------------------------------------------------------------

    def preprocess(self, input_file: str, output_path: str, **kwargs: Any) -> PreprocessorResult:
        """Preprocess a regulatory PDF.

        Args:
            input_file: Path to the PDF.
            output_path: Base path for output files (stem used as prefix).

        Returns:
            PreprocessorResult with two output files:
            - ``<stem>_clean.md``       — noise-stripped full text
            - ``<stem>_definitions.md`` — structured term-definition glossary
        """
        try:
            import pypdf  # noqa: PLC0415
        except ImportError as exc:
            raise ImportError(
                "pypdf is required by RegulatoryPDFPreprocessor. "
                "Run: uv add pypdf --package elt-llm-ingest"
            ) from exc

        input_path = Path(input_file).expanduser().resolve()
        output_base = Path(output_path).expanduser().resolve()
        output_base.parent.mkdir(parents=True, exist_ok=True)

        logger.info("%s: reading %s", type(self).__name__, input_path)

        reader = pypdf.PdfReader(str(input_path))
        total_pages = len(reader.pages)
        logger.info("  %d pages", total_pages)

        # ── Pass 1: extract and clean full text ───────────────────────────
        clean_sections: List[str] = []
        current_section: str = ""

        for page_idx, page in enumerate(reader.pages):
            raw = page.extract_text() or ""
            lines = raw.split("\n")

            new_section_header: Optional[str] = None
            content_lines: List[str] = []

            for line in lines:
                stripped = line.strip()
                if not stripped:
                    continue

                # Drop ALL lines that are section headers.
                # Headers repeat on every page (e.g. "8 - RULES OF THE ASSOCIATION")
                # and are high-frequency noise for BM25.  Track section changes so we
                # can emit a single ## marker per unique header.
                if self._HEADER_PATTERN.match(stripped):
                    if stripped != current_section:
                        new_section_header = stripped
                    continue

                # Drop noise-only lines
                if self._NOISE_LINES.match(stripped):
                    continue

                # Drop lines that are standalone page numbers or "NNN SECTION HEADER"
                # variants (page number prefix from PDF footer/TOC artifacts)
                if re.match(r"^\d{1,4}$", stripped):
                    continue
                if re.match(r"^\d{1,4}\s+\d+\s*[-–]", stripped):
                    continue

                content_lines.append(stripped)

            if new_section_header and new_section_header != current_section:
                # Flush previous section and start a new one
                if clean_sections:
                    clean_sections.append("")  # blank line between sections
                clean_sections.append(f"## {new_section_header}")
                current_section = new_section_header

            if content_lines:
                clean_sections.extend(content_lines)

        clean_md = "\n".join(clean_sections)

        clean_path = output_base.parent / f"{output_base.stem}_clean.md"
        clean_path.write_text(clean_md, encoding="utf-8")
        logger.info("  Written clean text: %s (%d chars)", clean_path, len(clean_md))

        # ── Pass 2: extract definitions glossary ──────────────────────────
        # Two independent extraction strategies are run and merged:
        #   "explicit"  — page ranges from def_sections config (highest confidence)
        #   "detected"  — pages auto-detected by means-pattern density
        #   "both"      — term found by both strategies (strongest signal)

        # store: term_lower → (canonical_term, defn)
        explicit_store: Dict[str, Tuple[str, str]] = {}
        detected_store: Dict[str, Tuple[str, str]] = {}

        means_pat = re.compile(
            r"([A-Z][A-Za-z0-9\s/\-'\"]{1,50}?)\s+means\s+(.+?)(?=;|(?=[A-Z][A-Za-z0-9\s/\-'\"]{1,50}?\s+means\s))",
            re.DOTALL,
        )
        new_def_line = re.compile(
            r"^([A-Z][A-Za-z0-9\s/'\"\-]{1,60}?)\s{2,}-\s+(.+)"
        )

        def _collect(store: Dict[str, Tuple[str, str]], term: str, defn: str) -> None:
            term = " ".join(term.split())
            defn = " ".join(defn.split()).strip(";").strip()
            if len(defn) < 10 or len(defn) > 1500:
                return
            if term.lower() not in store:
                store[term.lower()] = (term, defn)

        def _extract_page_range(
            page_indices: List[int],
            store: Dict[str, Tuple[str, str]],
        ) -> None:
            """Run Pattern A (means) and Pattern B (  -  ) over a list of pages."""
            # Pattern A: joined text
            joined_parts: List[str] = []
            for pg_idx in page_indices:
                text = reader.pages[pg_idx].extract_text() or ""
                lines = text.split("\n")
                lines = [l for l in lines if not self._HEADER_PATTERN.match(l.strip())]
                joined_parts.append(" ".join(lines))
            joined = " ".join(joined_parts)
            for m in means_pat.finditer(joined):
                _collect(store, m.group(1), m.group(2))

            # Pattern B: line-by-line accumulator (handles wrapped definitions)
            all_lines: List[str] = []
            for pg_idx in page_indices:
                raw = reader.pages[pg_idx].extract_text() or ""
                for ln in raw.split("\n"):
                    stripped = ln.strip()
                    if stripped and not self._HEADER_PATTERN.match(stripped):
                        all_lines.append(stripped)
            current_term: Optional[str] = None
            current_defn: List[str] = []
            for ln in all_lines:
                m = new_def_line.match(ln)
                if m:
                    if current_term:
                        _collect(store, current_term, " ".join(current_defn))
                    current_term = m.group(1).strip()
                    current_defn = [m.group(2).strip()]
                elif current_term and ln:
                    current_defn.append(ln)
            if current_term:
                _collect(store, current_term, " ".join(current_defn))

        # ── Auto-detect: scan all pages by means-pattern density ──────────
        if self.def_density_threshold > 0:
            density_pat = re.compile(
                r"[A-Z][A-Za-z0-9\s/\-'\"]{1,50}?\s+means\s+"
            )
            auto_pages: List[int] = []
            for pg_idx in range(total_pages):
                text = reader.pages[pg_idx].extract_text() or ""
                if len(density_pat.findall(text)) >= self.def_density_threshold:
                    auto_pages.append(pg_idx)
            logger.info("  Auto-detected %d definition pages (threshold=%d)",
                        len(auto_pages), self.def_density_threshold)
            if auto_pages:
                _extract_page_range(auto_pages, detected_store)

        # ── Explicit page ranges ───────────────────────────────────────────
        for page_start, page_end in self.def_sections:
            page_indices = list(range(
                min(page_start, total_pages),
                min(page_end, total_pages),
            ))
            _extract_page_range(page_indices, explicit_store)

        # ── Merge with confidence tagging ─────────────────────────────────
        # definitions: (term, defn, source)
        merged: List[Tuple[str, str, str]] = []
        for key, (term, defn) in explicit_store.items():
            source = "both" if key in detected_store else "explicit"
            merged.append((term, defn, source))
        for key, (term, defn) in detected_store.items():
            if key not in explicit_store:
                merged.append((term, defn, "detected"))

        logger.info(
            "  Definitions: %d explicit, %d detected, %d both, %d total unique",
            sum(1 for _, _, s in merged if s == "explicit"),
            sum(1 for _, _, s in merged if s == "detected"),
            sum(1 for _, _, s in merged if s == "both"),
            len(merged),
        )

        doc_title = Path(input_path).stem.replace("_", " ").replace("-", " ")
        def_lines: List[str] = [
            f"# {doc_title} — Defined Terms\n",
            f"The following terms are formally defined in {doc_title}.\n",
            "---\n",
        ]
        for term, defn, source in sorted(merged, key=lambda x: x[0].lower()):
            def_lines.append(f"\n## {term}\n")
            def_lines.append(
                f"**{self.definitions_label}** [source: {source}]: {term} means {defn}.\n"
            )

        def_md = "\n".join(def_lines)
        def_path = output_base.parent / f"{output_base.stem}_definitions.md"
        def_path.write_text(def_md, encoding="utf-8")
        logger.info("  Written definitions glossary: %s (%d terms)", def_path, len(merged))

        output_files = [str(clean_path), str(def_path)]
        return PreprocessorResult(
            original_file=str(input_path),
            output_files=output_files,
            success=True,
            message=(
                f"Processed {total_pages} pages; "
                f"{len(merged)} definitions extracted"
            ),
        )


class IdentityPreprocessor(BasePreprocessor):
    """Pass-through preprocessor that returns the original file.

    Used when no preprocessing is needed.
    """
    
    def preprocess(self, input_file: str, output_path: str, **kwargs: Any) -> PreprocessorResult:
        """Return the original file unchanged.
        
        Args:
            input_file: Path to the input file.
            output_path: Unused (kept for interface compatibility).
            **kwargs: Unused.
            
        Returns:
            PreprocessorResult with the original file path.
        """
        return PreprocessorResult(
            original_file=input_file,
            output_files=[input_file],
            success=True,
            message="No preprocessing applied"
        )


@dataclass
class PreprocessorConfig:
    """Preprocessor configuration.

    Attributes:
        module: Python module containing the preprocessor class.
        class_name: Name of the preprocessor class.
        output_format: Output format (e.g. 'markdown', 'json', 'both', 'split').
        output_suffix: Suffix for output files (default: '_processed').
        enabled: Whether preprocessing is enabled.
        collection_prefix: For split-mode preprocessors — prefix used to name
            each section's target collection (e.g. 'fa_leanix' → 'fa_leanix_agreements').
        extra_params: Additional keyword arguments forwarded as-is to the
            preprocessor constructor.  Use this to pass preprocessor-specific
            settings (e.g. ``def_sections``, ``model_name``) from YAML config
            without modifying PreprocessorConfig itself.
    """
    module: str
    class_name: str
    output_format: str = "markdown"
    output_suffix: str = "_processed"
    enabled: bool = True
    collection_prefix: Optional[str] = None
    extra_params: Dict[str, Any] = field(default_factory=dict)

    _KNOWN_KEYS = frozenset(
        {"module", "class", "output_format", "output_suffix", "enabled", "collection_prefix"}
    )

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "PreprocessorConfig":
        """Create PreprocessorConfig from a dictionary.

        Any keys not in the standard set are collected into ``extra_params``
        and forwarded to the preprocessor constructor, allowing per-class
        configuration (e.g. ``def_sections``, ``model_name``) to live in
        YAML without requiring changes to this class.
        """
        extra = {k: v for k, v in data.items() if k not in cls._KNOWN_KEYS}
        return cls(
            module=data.get("module", ""),
            class_name=data.get("class", ""),
            output_format=data.get("output_format", "markdown"),
            output_suffix=data.get("output_suffix", "_processed"),
            enabled=data.get("enabled", True),
            collection_prefix=data.get("collection_prefix"),
            extra_params=extra,
        )


def get_preprocessor(config: PreprocessorConfig) -> BasePreprocessor:
    """Factory function to create a preprocessor instance.
    
    Args:
        config: Preprocessor configuration.
        
    Returns:
        Preprocessor instance.
        
    Raises:
        ImportError: If the preprocessor module cannot be imported.
        AttributeError: If the preprocessor class doesn't exist.
    """
    if not config.enabled:
        return IdentityPreprocessor()
    
    module = __import__(config.module, fromlist=[config.class_name])
    preprocessor_class = getattr(module, config.class_name)
    return preprocessor_class(
        output_format=config.output_format,
        collection_prefix=config.collection_prefix,
        **config.extra_params,
    )


def preprocess_file(
    input_file: str,
    config: Optional[PreprocessorConfig],
    output_dir: Optional[str] = None
) -> List[str]:
    """Preprocess a file using the specified configuration.
    
    Args:
        input_file: Path to the input file.
        config: Preprocessor configuration. If None, returns original file.
        output_dir: Directory for output files. If None, uses input file's directory.
        
    Returns:
        List of output file paths (may be the original file if no preprocessing).
    """
    if config is None or not config.enabled:
        logger.debug("No preprocessing configured for: %s", input_file)
        return [input_file]
    
    try:
        preprocessor = get_preprocessor(config)
    except (ImportError, AttributeError) as e:
        logger.error("Failed to load preprocessor: %s", e)
        return [input_file]
    
    # Generate output path
    input_path = Path(input_file)
    if output_dir:
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        output_path = output_dir_path / f"{input_path.stem}{config.output_suffix}"
    else:
        output_path = input_path.parent / f"{input_path.stem}{config.output_suffix}"
    
    # Run preprocessor
    result = preprocessor.preprocess(str(input_file), str(output_path))
    
    if result.success:
        logger.info("Preprocessing successful: %s", result.message)
        return result.output_files
    else:
        logger.warning("Preprocessing failed (%s), using original file", result.message)
        return [input_file]
